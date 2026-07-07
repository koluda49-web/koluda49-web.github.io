#!/usr/bin/env python3
"""
Полный пайплайн: CRM -> МТС VATS -> Google Таблица

КОМАНДЫ:

1. Выгрузить заказы из CRM и подготовить файл номеров для МТС:
   python scrape_crm.py crm

2. Запустить автоинформирование в МТС (создаёт НОВОЕ чистое задание с нуля,
   настраивает синтез речи и сбор ответов, загружает свежий список номеров,
   сохраняет — обзвон стартует сам):
   python scrape_crm.py call

   Запасной вариант (копирование старого задания — может звонить и на
   старые номера шаблона, используй только если создание с нуля не работает):
   python scrape_crm.py call --copy [ССЫЛКА_НА_СУЩЕСТВУЮЩУЮ_КОПИЮ]

3. Скачать отчёт по завершённому заданию и сразу объединить с заказами CRM:
   python scrape_crm.py fetch_report ССЫЛКА_НА_ЗАДАНИЕ

4. (Вручную) Объединить уже скачанный отчёт МТС с заказами из CRM:
   python scrape_crm.py merge ОТЧЕТ_МТС.csv

Установка:
    pip install selenium beautifulsoup4 webdriver-manager openpyxl pandas requests
"""

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv
import time
import sys
import re
from datetime import datetime

import requests

CONFIG_FILE = "config.txt"


def load_config() -> dict:
    """Читает config.txt (логин/пароль МТС) из той же папки, что и скрипт."""
    config = {}
    if not os.path.exists(CONFIG_FILE):
        return config
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            config[key.strip()] = value.strip()
    return config


def try_mts_login(driver):
    """
    Если открылась страница входа МТС — пытается автоматически ввести
    логин/пароль из config.txt. Если конфига нет или поля не найдены,
    просто ничего не делает (пользователь войдёт вручную как раньше).
    """
    config = load_config()
    login = config.get("MTS_LOGIN")
    password = config.get("MTS_PASSWORD")

    if not login or not password or password == "ВСТАВЬ_СЮДА_ПАРОЛЬ":
        return False  # конфиг не настроен — пропускаем автологин

    try:
        # Ищем поле логина: на форме входа МТС это первый input[type='text']
        # (без name='login'/'email' — просто обычное текстовое поле)
        login_field = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((
                By.CSS_SELECTOR,
                "input[type='text']"
            ))
        )
        password_field = driver.find_element(
            By.CSS_SELECTOR,
            "input[type='password']"
        )

        login_field.clear()
        login_field.send_keys(login)
        password_field.clear()
        password_field.send_keys(password)

        # Ищем кнопку входа
        try:
            submit_btn = driver.find_element(
                By.XPATH,
                "//button[@type='submit'] | //button[contains(.,'Войти')] | //button[contains(.,'Вход')]"
            )
            submit_btn.click()
        except Exception:
            from selenium.webdriver.common.keys import Keys
            password_field.send_keys(Keys.ENTER)

        print("   🔑 Логин выполнен автоматически (из config.txt)")
        time.sleep(2)
        return True

    except Exception:
        return False  # форма входа не найдена — значит уже авторизован или другая страница

CRM_OUTPUT_FILE   = "crm_orders.xlsx"      # выгрузка из CRM (с привязкой к заказу)
MTS_NUMBERS_FILE  = "mts_numbers.csv"      # файл со списком номеров для загрузки в МТС
MERGED_OUTPUT_FILE = "crm_with_results.xlsx"  # итоговый файл с результатами обзвона
LAST_TASK_URL_FILE  = "last_task_url.txt"   # ссылка на последнее созданное задание МТС
LAST_TASK_NAME_FILE = "last_task_name.txt"  # название последнего задания (для поиска в списке)

# Постоянная ссылка на задание-шаблон в МТС VATS (используется командой call --copy,
# оставлена как запасной вариант на случай если потребуется копирование).
MTS_TEMPLATE_TASK_URL = "https://193543133-18519208.vats.mts.by/#/autocaller/tasks/3a4156a6-e990-4191-ad23-074cb68846af"
MTS_TASKS_LIST_URL = "https://193543133-18519208.vats.mts.by/#/autocaller/tasks"
MTS_NEW_TASK_URL = "https://193543133-18519208.vats.mts.by/#/autocaller/tasks/new"

# Текст сообщения для синтеза речи и название опроса сбора ответов —
# используются при создании нового задания с нуля.
MTS_MESSAGE_TEXT = (
    "!!!-Здравствуйте! Вас беспокоит робот логист, - вы оставили заказ в "
    "интернет магазине Зикмес. Мы хотим подтвердить доставку товара. "
    "Чтобы выбрать ответ, нажмите кнопку на телефоне: -"
)
MTS_SURVEY_NAME = "Итог"

import os
DOWNLOADS_DIR = os.path.join(os.path.expanduser("~"), "Downloads")

# URL веб-приложения Google Apps Script (созданного через "Развернуть -> Веб-приложение")
GOOGLE_SHEET_WEBHOOK = "https://script.google.com/macros/s/AKfycbyLoeIWSWVnlVME4xbbYSyFm_mjVWM7gpjUvGHDbp9mG4lBWny4R9X9W878j6gC03fucw/exec"


def send_to_google_sheet(action: str, rows: list[dict]):
    """Отправляет данные в Google Таблицу через Apps Script веб-приложение."""
    try:
        resp = requests.post(
            GOOGLE_SHEET_WEBHOOK,
            json={"action": action, "rows": rows},
            timeout=60,
        )
        result = resp.json()
        if result.get("success"):
            print(f"  ☁️  Google Таблица обновлена: {result.get('message')}")
        else:
            print(f"  ⚠️  Ошибка записи в Google Таблицу: {result.get('message')}")
    except Exception as e:
        print(f"  ⚠️  Не удалось связаться с Google Таблицей: {e}")

ROUTES = [
    {"name": "б",                  "url": "https://a.ok-crm.com/route/cabinet-calls?route_id=17"},
    {"name": "а",                  "url": "https://a.ok-crm.com/route/cabinet-calls?route_id=30"},
    {"name": "в",                  "url": "https://a.ok-crm.com/route/cabinet-calls?route_id=58"},
    {"name": "минск",              "url": "https://a.ok-crm.com/route/cabinet-calls?route_id=7768"},
    {"name": "2 не трогать тест",  "url": "https://a.ok-crm.com/route/cabinet-calls?route_id=29116"},
]


# ============================================================
#  ЧАСТЬ 1 — Выгрузка из CRM
# ============================================================

def create_driver():
    try:
        from webdriver_manager.chrome import ChromeDriverManager
        service = Service(ChromeDriverManager().install())
    except ImportError:
        service = Service()

    options = webdriver.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    # Сохраняем профиль Chrome в папку рядом со скриптом.
    # Это позволяет не вводить логин/пароль Google при каждом запуске —
    # куки и сессия сохраняются между запусками в этой папке.
    profile_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chrome_profile")
    os.makedirs(profile_dir, exist_ok=True)
    options.add_argument(f"--user-data-dir={profile_dir}")

    # Отключаем всплывающее окно "Сохранить пароль?" — оно может перекрывать
    # страницу и мешать Selenium кликать по элементам сразу после автологина.
    prefs = {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
        "profile.password_manager_leak_detection": False,
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(service=service, options=options)
    driver.maximize_window()
    return driver


def wait_for_manual_login(driver):
    driver.get(ROUTES[0]["url"])
    time.sleep(3)

    # Проверяем — если уже залогинены (сессия сохранена в chrome_profile),
    # таблица с заказами загрузится сама и Enter вводить не нужно.
    try:
        WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "tr.grid-order-widget, table.items"))
        )
        print("  ✅ Сессия CRM активна — вход не требуется.")
        return
    except Exception:
        pass

    # Сессия не найдена — просим залогиниться вручную (первый раз или после сброса)
    print()
    print("=" * 55)
    print("  Залогинься через Google в открывшемся браузере.")
    print("  После первого входа логин сохранится и больше")
    print("  вводить его не понадобится.")
    print("  Дождись пока загрузится таблица с заказами.")
    print("  Нажми Enter здесь когда увидишь таблицу.")
    print("=" * 55)
    input("  -> Нажми Enter: ")
    print()


def scroll_and_collect(driver) -> list[dict]:
    print("  Жду загрузки строк...")
    try:
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "tr.grid-order-widget"))
        )
    except Exception:
        print("  Строки не появились за 20 сек, пробую парсить...")

    collected = {}
    scroll_attempts = 0
    max_attempts = 30
    last_count = 0

    while scroll_attempts < max_attempts:
        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")
        order_rows = soup.find_all("tr", class_="grid-order-widget")

        for tr in order_rows:
            tds = tr.find_all("td")
            if len(tds) < 6:
                continue
            row_id = clean(tds[1].get_text())
            if not row_id:
                continue
            if row_id not in collected:
                # Статус — предпоследняя колонка перед «Действия».
                # Порядок колонок: чекбокс(0), ID(1), товары(2), цены(3),
                # телефон(4), адрес(5), пометка(6), доставка с(7), дилер(8), статус(9), действия(10)
                status = clean(tds[9].get_text()) if len(tds) > 9 else ""
                collected[row_id] = {
                    "ID":              row_id,
                    "Товары":          clean(tds[2].get_text()),
                    "Цены":            clean(tds[3].get_text()),
                    "Телефон_сырой":   clean(tds[4].get_text()),
                    "Адрес доставки":  clean(tds[5].get_text()),
                    "Статус_CRM":      status,
                }

        current_count = len(collected)
        print(f"  Собрано: {current_count}", end="\r")

        if current_count == last_count:
            scroll_attempts += 1
        else:
            scroll_attempts = 0
            last_count = current_count

        try:
            driver.execute_script("""
                var container = document.querySelector('.kv-grid-container')
                             || document.querySelector('.table-responsive')
                             || document.querySelector('.grid-view')
                             || document.body;
                container.scrollTop += 500;
                window.scrollBy(0, 300);
            """)
        except Exception:
            pass

        time.sleep(0.5)

    print()
    return list(collected.values())


def clean(text: str) -> str:
    return " ".join(text.split()).strip()


def extract_phone(raw: str) -> str:
    """
    Извлекает первый номер телефона из строки и нормализует в формат 375XXXXXXXXX.
    Примеры входа: '+375 44 709-98-51, +375 2235 55-437' -> '375447099851'
                   'Иванов Иван+375 29 698-18-39'          -> '375296981839'
    """
    # Ищем все последовательности похожие на номер с +375 или 375 или 80
    matches = re.findall(r'(?:\+?375|80)[\d\s\-\(\)]{7,}', raw)
    if not matches:
        return ""

    first = matches[0]
    digits = re.sub(r'\D', '', first)  # убираем всё кроме цифр

    # Нормализуем к формату 375XXXXXXXXX (12 цифр)
    if digits.startswith("00375"):
        digits = digits[2:]
    if digits.startswith("80"):
        digits = "375" + digits[2:]
    if digits.startswith("375") and len(digits) == 12:
        return digits
    if digits.startswith("375") and len(digits) > 12:
        return digits[:12]

    return digits  # вернём как есть если не подошло под правило, для проверки вручную


def write_sheet(ws, rows: list[dict]):
    headers = ["ID", "Телефон клиента", "Адрес доставки", "Открыть в CRM"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        phone = extract_phone(row["Телефон_сырой"])
        ws.cell(row=row_idx, column=1, value=row["ID"])
        ws.cell(row=row_idx, column=2, value=phone)
        ws.cell(row=row_idx, column=3, value=row["Адрес доставки"])

        order_id = row["ID"]
        if order_id:
            link_cell = ws.cell(row=row_idx, column=4, value="Открыть →")
            link_cell.hyperlink = f"https://a.ok-crm.com/order/update/{order_id}"
            link_cell.font = Font(color="0563C1", underline="single")
            link_cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 14  # ID
    ws.column_dimensions["B"].width = 18  # Телефон
    ws.column_dimensions["C"].width = 45  # Адрес
    ws.column_dimensions["D"].width = 18  # Ссылка


def cmd_crm():
    """Команда: выгрузить из CRM и подготовить файл номеров для МТС."""

    # Выбор маршрута перед запуском браузера
    print()
    print("=" * 55)
    print("  Выбери маршруты для выгрузки:")
    print("=" * 55)
    print("  0. Все маршруты")
    for i, route in enumerate(ROUTES, 1):
        print(f"  {i}. {route['name']}")
    print("=" * 55)
    print("  Можно выбрать несколько через запятую: 1,3")
    while True:
        choice = input("  Введи цифру (или несколько через запятую): ").strip()
        if choice == "0":
            selected_routes = ROUTES
            print(f"  Выбрано: все маршруты ({len(ROUTES)} шт.)")
            break
        parts = [p.strip() for p in choice.split(",")]
        if all(p.isdigit() and 1 <= int(p) <= len(ROUTES) for p in parts):
            selected_routes = [ROUTES[int(p) - 1] for p in parts]
            names = ", ".join(r["name"] for r in selected_routes)
            print(f"  Выбрано: {names}")
            break
        print(f"  Неверный выбор, введи цифры от 0 до {len(ROUTES)} через запятую")
    print()

    driver = create_driver()
    all_rows = []  # все строки со всех выбранных маршрутов, для файла МТС

    try:
        wait_for_manual_login(driver)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        total = 0
        for route in selected_routes:
            print(f"\n[Маршрут: {route['name']}] {route['url']}")
            driver.get(route["url"])
            time.sleep(2)

            rows = scroll_and_collect(driver)
            print(f"  Собрано всего: {len(rows)} записей")

            # Фильтруем — берём только заказы со статусом «В наличии»
            rows_filtered = [r for r in rows if r.get("Статус_CRM", "").strip().lower() == "в наличии"]
            skipped = len(rows) - len(rows_filtered)
            print(f"  После фильтра «В наличии»: {len(rows_filtered)} записей (пропущено {skipped} с другим статусом)")

            # Защита: если строки были, но после фильтра ноль — возможно
            # индекс колонки статуса сдвинулся (CRM поменяла разметку).
            # Показываем примеры найденных статусов и спрашиваем продолжать ли.
            if len(rows) > 0 and len(rows_filtered) == 0:
                sample_statuses = list({r.get("Статус_CRM", "(пусто)") for r in rows[:10]})
                print(f"\n  ⚠️  Ни одна строка не прошла фильтр «В наличии»!")
                print(f"  Примеры статусов в таблице: {sample_statuses}")
                print(f"  Возможно, разметка CRM изменилась и колонка статуса сдвинулась.")
                ans = input("  Продолжить без фильтра (загрузить ВСЕ статусы)? [д/н]: ").strip().lower()
                if ans in ("д", "y", "yes", "да"):
                    rows_filtered = rows
                    print("  Продолжаю без фильтра по статусу.")
                else:
                    print("  Пропускаю этот маршрут.")
                    rows_filtered = []

            rows = rows_filtered
            total += len(rows)

            ws = wb.create_sheet(title=route["name"])
            write_sheet(ws, rows)

            for row in rows:
                row["Маршрут"] = route["name"]
            all_rows.extend(rows)

        ws_info = wb.create_sheet(title="Инфо")
        ws_info["A1"] = "Последнее обновление:"
        ws_info["B1"] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        ws_info["A2"] = "Всего записей:"
        ws_info["B2"] = total

        wb.save(CRM_OUTPUT_FILE)
        print(f"\n✅ CRM-выгрузка сохранена -> {CRM_OUTPUT_FILE} ({total} записей)")

        # Отправляем все строки в Google Таблицу (лист "Заказы")
        print("\n☁️  Отправляю данные в Google Таблицу...")
        gsheet_rows = []
        for row in all_rows:
            phone = extract_phone(row["Телефон_сырой"])
            gsheet_rows.append({
                "ID":              row["ID"],
                "Телефон клиента": phone,
                "Адрес доставки":  row["Адрес доставки"],
                "Маршрут":         row["Маршрут"],
            })
        send_to_google_sheet("upload_crm", gsheet_rows)

        # Готовим файл номеров для МТС: только номера, без заголовка, уникальные, формат 375XXXXXXXXX
        phones = set()
        for row in all_rows:
            phone = extract_phone(row["Телефон_сырой"])
            if re.fullmatch(r"375\d{9}", phone):
                phones.add(phone)
            else:
                print(f"  ⚠️  Пропущен некорректный номер (заказ {row['ID']}): '{row['Телефон_сырой']}'")

        # Сохраняем .xlsx-версию с числовым форматом (0 знаков после запятой),
        # чтобы при открытии в Excel номера не превращались в 3,75E+11
        mts_xlsx_file = MTS_NUMBERS_FILE.replace(".csv", ".xlsx")
        wb_mts = openpyxl.Workbook()
        ws_mts = wb_mts.active
        ws_mts.title = "Номера"
        for i, phone in enumerate(sorted(phones), 1):
            cell = ws_mts.cell(row=i, column=1, value=int(phone))
            cell.number_format = "0"  # числовой формат, без дробной части
        ws_mts.column_dimensions["A"].width = 18
        wb_mts.save(mts_xlsx_file)

        print(f"✅ Файл для МТС сохранён -> {mts_xlsx_file} ({len(phones)} уникальных номеров)")
        print("\nТеперь загрузи этот файл в задание автоинформирования в МТС VATS.")

    finally:
        driver.quit()
        print("Браузер закрыт.")


# ============================================================
#  ЧАСТЬ 2 — Объединение отчёта МТС с заказами CRM
# ============================================================

def load_crm_data() -> dict:
    """Загружает все заказы из CRM-выгрузки в словарь {телефон: [список заказов]}."""
    if not openpyxl.load_workbook:
        pass

    wb = openpyxl.load_workbook(CRM_OUTPUT_FILE, data_only=True)
    phone_to_orders = {}

    for sheet_name in wb.sheetnames:
        if sheet_name == "Инфо":
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        try:
            id_col = headers.index("ID")
            goods_col = headers.index("Товары")
            price_col = headers.index("Цены")
            phone_col = headers.index("Телефон клиента")
            addr_col = headers.index("Адрес доставки")
        except ValueError:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[phone_col]:
                continue
            phone_raw = str(row[phone_col]).strip()
            # Убираем .0 если Excel прочитал номер как число (375295580089.0 -> 375295580089)
            phone_raw = re.sub(r'\.0+$', '', phone_raw)
            phone = normalize_phone(phone_raw)
            if not phone:
                continue
            order = {
                "ID":              row[id_col],
                "Товары":          row[goods_col],
                "Цены":            row[price_col],
                "Телефон клиента": phone,
                "Адрес доставки":  row[addr_col],
                "Маршрут":         sheet_name,
            }
            phone_to_orders.setdefault(phone, []).append(order)

    return phone_to_orders


ANSWER_LABELS = {
    "1": "Клиент готов принять",
    "2": "Клиент не готов принять",
    "3": "Связь с оператором",
}


def load_mts_report(filepath: str) -> dict:
    """
    Загружает отчёт МТС (CSV, новый .xlsx или старый .xls) в формате:
    Номер | Часовой пояс | Статус | Количество попыток | Результат последнего вызова |
    Длительность прослушивания | Аудиофайл | Ответ

    Возвращает {телефон: результат}
    """
    phone_to_result = {}
    lower = filepath.lower()

    if lower.endswith(".csv"):
        with open(filepath, "r", encoding="utf-8-sig") as f:
            sample = f.read(2048)
            f.seek(0)
            delimiter = ";" if sample.count(";") > sample.count(",") else ","
            reader = csv.DictReader(f, delimiter=delimiter)
            for row in reader:
                _store_mts_row(row, phone_to_result)
        return phone_to_result

    # Пытаемся открыть как нормальный .xlsx через openpyxl
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            row_dict = dict(zip(headers, row))
            _store_mts_row(row_dict, phone_to_result)
        return phone_to_result
    except Exception as e:
        print(f"  ⚠️  Не удалось открыть как .xlsx ({e})")
        print("  Пробую открыть как старый формат .xls / HTML-таблицу...")

    # Запасной вариант: файл на самом деле .xls (старый формат) или HTML-таблица
    # с расширением .xlsx — пробуем через pandas, который умеет и то и другое.
    import pandas as pd
    df = None
    for engine in (None, "xlrd", "html"):
        try:
            if engine == "html":
                tables = pd.read_html(filepath)
                df = tables[0]
            else:
                df = pd.read_excel(filepath, engine=engine) if engine else pd.read_excel(filepath)
            break
        except Exception:
            continue

    if df is None:
        raise RuntimeError(
            f"Не удалось прочитать файл отчёта '{filepath}' ни одним способом.\n"
            f"Попробуй скачать отчёт в формате CSV из МТС VATS вместо Excel."
        )

    df.columns = [str(c).strip() for c in df.columns]
    for _, row in df.iterrows():
        row_dict = row.to_dict()
        _store_mts_row(row_dict, phone_to_result)

    return phone_to_result


def _store_mts_row(row: dict, phone_to_result: dict):
    """Извлекает данные из одной строки отчёта МТС и сохраняет в словарь по номеру."""
    phone_raw = str(row.get("Номер", "")).strip()
    phone = normalize_phone(phone_raw)
    if not phone:
        return

    status = str(row.get("Статус", "")).strip()
    answer_raw = row.get("Ответ", "")
    answer_str = str(answer_raw).strip() if answer_raw is not None else ""

    answer_label = ANSWER_LABELS.get(answer_str, "")
    if status.lower() != "успешно":
        result_text = f"Звонок не успешен ({status})" if status else "Нет данных"
    elif answer_label:
        result_text = answer_label
    elif answer_str:
        result_text = f"Ответ: {answer_str}"
    else:
        result_text = "Прослушано, без ответа"

    phone_to_result[phone] = result_text


def normalize_phone(raw: str) -> str:
    digits = re.sub(r'\D', '', raw)
    if digits.startswith("80"):
        digits = "375" + digits[2:]
    if digits.startswith("00375"):
        digits = digits[2:]
    if not digits.startswith("375"):
        return ""
    return digits[:12]


def cmd_merge(report_path: str):
    """Команда: объединить отчёт МТС с заказами CRM."""
    print(f"📥 Загружаю заказы из {CRM_OUTPUT_FILE}...")
    phone_to_orders = load_crm_data()
    total_orders = sum(len(v) for v in phone_to_orders.values())
    print(f"   Загружено заказов: {total_orders} (уникальных номеров: {len(phone_to_orders)})")

    print(f"📥 Загружаю отчёт МТС из {report_path}...")
    phone_to_result = load_mts_report(report_path)
    print(f"   Загружено результатов: {len(phone_to_result)}")

    # ДИАГНОСТИКА: показываем первые 3 номера с обеих сторон для сравнения
    crm_phones = list(phone_to_orders.keys())[:3]
    mts_phones = list(phone_to_result.keys())[:3]
    print(f"   📞 Примеры номеров из CRM:  {crm_phones}")
    print(f"   📞 Примеры номеров из МТС:  {mts_phones}")
    # Проверяем пересечение
    common = set(phone_to_orders.keys()) & set(phone_to_result.keys())
    print(f"   ✅ Совпало номеров: {len(common)} из {len(phone_to_orders)}")

    # Собираем итоговую таблицу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"

    headers = ["ID", "Телефон клиента", "Адрес доставки", "Маршрут", "Результат обзвона", "Открыть заказ в CRM"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Цвета строк по результату обзвона
    fill_green  = PatternFill("solid", fgColor="C6EFCE")  # Клиент готов
    fill_red    = PatternFill("solid", fgColor="FFC7CE")  # Не готов
    fill_yellow = PatternFill("solid", fgColor="FFEB9C")  # Оператор / без ответа
    fill_none   = PatternFill("solid", fgColor="F2F2F2")  # Нет данных

    def result_fill(result_text):
        if "готов принять" in result_text and "не" not in result_text.lower()[:10]:
            return fill_green
        if "не готов" in result_text:
            return fill_red
        if result_text in ("Нет данных / не дозвонились",) or result_text.startswith("Звонок не успешен"):
            return fill_none
        return fill_yellow

    row_idx = 2
    matched = 0
    gsheet_results = []
    for phone, orders in phone_to_orders.items():
        result = phone_to_result.get(phone, "Нет данных / не дозвонились")
        if phone in phone_to_result:
            matched += 1
        for order in orders:
            order_id = str(order["ID"]).replace(".0", "").strip()
            crm_url = f"https://a.ok-crm.com/order/update/{order_id}"
            row_fill = result_fill(result)

            for col, value in enumerate([
                order_id,
                order["Телефон клиента"],
                order["Адрес доставки"],
                order["Маршрут"],
                result,
            ], 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill

            # Колонка 6 — кликабельная ссылка на карточку заказа в CRM
            link_cell = ws.cell(row=row_idx, column=6, value="Открыть →")
            link_cell.hyperlink = crm_url
            link_cell.font = Font(color="0563C1", underline="single")
            link_cell.fill = row_fill
            link_cell.alignment = Alignment(horizontal="center")

            row_idx += 1

            gsheet_results.append({
                "ID":               order_id,
                "Телефон клиента":  order["Телефон клиента"],
                "Адрес доставки":   order["Адрес доставки"],
                "Маршрут":          order["Маршрут"],
                "Результат обзвона": result,
            })

    ws.column_dimensions["A"].width = 14  # ID — немного шире
    ws.column_dimensions["B"].width = 18  # Телефон
    ws.column_dimensions["C"].width = 45  # Адрес доставки
    ws.column_dimensions["D"].width = 12  # Маршрут
    ws.column_dimensions["E"].width = 30  # Результат обзвона
    ws.column_dimensions["F"].width = 18  # Ссылка

    wb.save(MERGED_OUTPUT_FILE)
    print(f"\n✅ Готово! -> {MERGED_OUTPUT_FILE}")
    print(f"   Сопоставлено по номеру: {matched} из {len(phone_to_orders)}")

    print("\n☁️  Отправляю результаты в Google Таблицу...")
    send_to_google_sheet("upload_results", gsheet_results)
    print("   Открой свою Google Таблицу — лист 'Результаты обзвона' обновлён.")

    # Удаляем файл отчёта МТС после успешного merge — он больше не нужен,
    # и чтобы не засорять папку C:\obzvon старыми файлами с uuid-именами.
    try:
        if os.path.exists(report_path) and os.path.abspath(report_path) != os.path.abspath(MERGED_OUTPUT_FILE):
            os.remove(report_path)
            print(f"   🗑️  Временный файл отчёта удалён: {os.path.basename(report_path)}")
    except Exception as e:
        print(f"   ⚠️  Не удалось удалить временный файл отчёта: {e}")


# ============================================================
#  ЧАСТЬ 3 — Автоматический запуск обзвона в МТС
# ============================================================

def cmd_call_copy(existing_copy_url: str = None):
    """
    [ЗАПАСНОЙ ВАРИАНТ] Открыть задание-шаблон в МТС, скопировать его, включить
    информирование, загрузить файл номеров и сохранить.

    ВНИМАНИЕ: МТС добавляет номера к уже существующим в задании, а не заменяет
    их — поэтому при копировании старые номера шаблона тоже получат звонок.
    Используй основную команду 'call' (cmd_call_new), которая создаёт чистое
    задание с нуля и не имеет этой проблемы.
    """
    mts_xlsx_file = MTS_NUMBERS_FILE.replace(".csv", ".xlsx")
    if not os.path.exists(mts_xlsx_file):
        print(f"❌ Файл {mts_xlsx_file} не найден. Сначала запусти: python scrape_crm.py crm")
        sys.exit(1)

    # Selenium принимает только абсолютный путь для send_keys в input[type=file]
    mts_xlsx_abspath = os.path.abspath(mts_xlsx_file)

    driver = create_driver()
    try:
        if existing_copy_url:
            print(f"🌐 Открываю уже существующую копию задания...")
            driver.get(existing_copy_url)
        else:
            print(f"🌐 Открываю задание-шаблон в МТС...")
            driver.get(MTS_TEMPLATE_TASK_URL)

        time.sleep(2)
        logged_in_automatically = try_mts_login(driver)

        if logged_in_automatically:
            # После автологина страница может перезагружаться/редиректить —
            # ждём явно, чтобы driver не работал со старым DOM.
            time.sleep(2)
            try:
                WebDriverWait(driver, 15).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
            except Exception:
                pass
            # Если после логина мы не на странице нужного задания — переоткрываем её явно
            target_url = existing_copy_url or MTS_TEMPLATE_TASK_URL
            if target_url.split("#")[-1] not in driver.current_url:
                driver.get(target_url)
                time.sleep(2)

        print()
        print("=" * 55)
        if logged_in_automatically:
            print("  Вход выполнен автоматически.")
        else:
            print("  Если МТС попросит войти — авторизуйся в открывшемся окне.")
        print("  Дождись пока полностью загрузится страница задания.")
        print("  Нажми Enter здесь, когда увидишь форму задания.")
        print("=" * 55)
        input("  -> Нажми Enter: ")
        print()

        windows_before = set(driver.window_handles)

        if not existing_copy_url:
            # Находим и нажимаем кнопку "Скопировать задание"
            print("📋 Копирую задание...")
            print("   Обновляю страницу для надёжности перед поиском кнопки...")
            driver.refresh()
            time.sleep(3)
            print("   (жду до 10 сек появления кнопки)")
            copy_btn = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((
                    By.XPATH, "//button[contains(., 'Скопировать задание')]"
                ))
            )
            # Прокручиваем к кнопке, чтобы её не перекрывали другие элементы (тултипы, шапка)
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", copy_btn)
            time.sleep(0.5)
            try:
                copy_btn.click()
            except Exception:
                # Обычный клик может быть перехвачен другим элементом — кликаем через JS
                driver.execute_script("arguments[0].click();", copy_btn)
            time.sleep(2)

            # Если копирование открыло новую вкладку/окно — переключаемся на неё
            windows_after = set(driver.window_handles)
            new_windows = windows_after - windows_before
            if new_windows:
                driver.switch_to.window(new_windows.pop())
                print("   (открылась новая вкладка — переключился на неё)")

        # Ждём загрузки формы задания (копии или уже открытой существующей)
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Включить информирование')]"))
        )
        time.sleep(1)  # дополнительная пауза на отрисовку
        print(f"✅ Задание готово к настройке: {driver.current_url}")

        # Включаем переключатель "Включить информирование" (если ещё не включён)
        print("🔘 Включаю информирование...")
        try:
            toggle = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "label.itl-switch input[type='checkbox']"))
            )
        except Exception:
            # Запасной вариант: ищем любой переключатель рядом с текстом "Включить информирование"
            toggle = driver.find_element(
                By.XPATH,
                "//*[contains(text(),'Включить информирование')]/following::input[@type='checkbox'][1]"
            )

        if not toggle.is_selected():
            try:
                label = driver.find_element(By.CSS_SELECTOR, "label.itl-switch")
                label.click()
            except Exception:
                # Клик через JS как запасной вариант, если обычный клик перехватывается другим элементом
                driver.execute_script("arguments[0].click();", toggle)
            time.sleep(0.5)

        # Загружаем новый файл номеров
        print(f"📤 Загружаю файл номеров: {mts_xlsx_abspath}")
        file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file'].file-input")
        file_input.send_keys(mts_xlsx_abspath)

        # ВАЖНО: после загрузки файла появляется модальное окно подтверждения
        # "Добавление номеров" с текстом вида "будет добавлен N уникальный номер"
        # и кнопкой "Загрузить список". Без явного клика по ней новый список
        # НЕ применяется к заданию, даже если файл был принят сервером.
        time.sleep(2)  # даём окну время появиться после send_keys
        print("⏳ Жду модальное окно подтверждения номеров (до 25 сек)...")
        try:
            confirm_btn = WebDriverWait(driver, 25).until(
                EC.element_to_be_clickable((
                    By.XPATH, "//button[contains(., 'Загрузить список')]"
                ))
            )
            # Пытаемся прочитать текст модального окна для лога (сколько номеров будет добавлено)
            try:
                modal_text = driver.find_element(
                    By.XPATH, "//*[contains(text(),'будет добавлен')]"
                ).text
                print(f"   Модальное окно: «{modal_text}»")
            except Exception:
                pass

            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", confirm_btn)
            time.sleep(0.3)
            try:
                confirm_btn.click()
            except Exception:
                driver.execute_script("arguments[0].click();", confirm_btn)
            print("   ✅ Подтвердил загрузку списка номеров")
            time.sleep(2)
        except Exception:
            # Даём ещё один шанс — возможно окно появилось чуть позже таймаута
            time.sleep(3)
            try:
                confirm_btn = driver.find_element(By.XPATH, "//button[contains(., 'Загрузить список')]")
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", confirm_btn)
                driver.execute_script("arguments[0].click();", confirm_btn)
                print("   ✅ Подтвердил загрузку списка номеров (со второй попытки)")
                time.sleep(2)
            except Exception:
                print("   ⚠️  Модальное окно подтверждения не появилось или не найдено.")
                print("   ⚠️  ОБЯЗАТЕЛЬНО посмотри на экран браузера прямо сейчас:")
                print("   ⚠️  если видишь окно 'Добавление номеров' — нажми в нём")
                print("   ⚠️  кнопку 'Загрузить список' САМ, и только потом жми Enter здесь.")
                input("   -> Нажми Enter здесь ПОСЛЕ клика в браузере (или если окна не было): ")

        # Сохраняем задание — это и запускает обзвон.
        # Финальная проверка делается визуально пользователем, так как текст с
        # количеством номеров на странице ненадёжно ловится программно.
        print("\n" + "=" * 55)
        print("  ПРОВЕРЬ НА ЭКРАНЕ БРАУЗЕРА:")
        print("  правильное ли количество номеров указано в разделе")
        print("  «Список для информирования»?")
        print("=" * 55)
        print("  Если всё верно — нажми Enter, чтобы сохранить и запустить обзвон.")
        print("  Если что-то не так — закрой окно браузера вместо Enter.")
        input("  -> Нажми Enter: ")

        print("💾 Сохраняю задание (это запускает обзвон)...")
        save_btn = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//button[contains(., 'Сохранить')]"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", save_btn)
        time.sleep(0.5)
        try:
            save_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", save_btn)
        time.sleep(3)

        print(f"\n✅ Обзвон запущен!")
        print(f"   Ссылка на задание: {driver.current_url}")
        print(f"\n   Сохрани эту ссылку — она понадобится команде fetch_report.")
        print(f"   Обычно обзвон полностью завершается (с учётом")
        print(f"   перезвонов) примерно за 15 минут.")

    finally:
        driver.quit()
        print("Браузер закрыт.")


def cmd_call_new(external_driver=None):
    """
    Команда: создать НОВОЕ чистое задание автоинформирования с нуля
    (вместо копирования старого), настроить синтез речи, сбор ответов,
    загрузить свежий файл номеров и сохранить.

    Это основной способ запуска обзвона — он не имеет проблемы с
    накоплением старых номеров, в отличие от копирования задания.
    """
    mts_xlsx_file = MTS_NUMBERS_FILE.replace(".csv", ".xlsx")
    if not os.path.exists(mts_xlsx_file):
        print(f"❌ Файл {mts_xlsx_file} не найден. Сначала запусти: python scrape_crm.py crm")
        sys.exit(1)

    mts_xlsx_abspath = os.path.abspath(mts_xlsx_file)

    # Если передан внешний браузер (полный цикл) — используем его и не закрываем.
    # Если нет — создаём свой и закрываем в finally.
    own_driver = external_driver is None
    driver = external_driver if external_driver else create_driver()
    try:
        print("🌐 Открываю список заданий в МТС...")
        driver.get(MTS_TASKS_LIST_URL)
        time.sleep(2)

        logged_in_automatically = try_mts_login(driver)
        if logged_in_automatically:
            time.sleep(2)
            try:
                WebDriverWait(driver, 15).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
            except Exception:
                pass
            if "tasks" not in driver.current_url:
                driver.get(MTS_TASKS_LIST_URL)
                time.sleep(2)

        print()
        print("=" * 55)
        if logged_in_automatically:
            print("  Вход выполнен автоматически.")
        else:
            print("  Если МТС попросит войти — авторизуйся в открывшемся окне.")
        print("  Дождись пока полностью загрузится список заданий.")
        print("  Нажми Enter здесь, когда увидишь кнопку «Добавить задание».")
        print("=" * 55)
        input("  -> Нажми Enter: ")
        print()

        print("➕ Нажимаю «Добавить задание»...")
        add_btn = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Добавить задание')]"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", add_btn)
        time.sleep(0.3)
        try:
            add_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", add_btn)
        time.sleep(2)

        print("⏳ Жду загрузки формы создания задания...")
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Основные данные')]"))
        )
        time.sleep(1)
        task_form_url = driver.current_url
        print(f"✅ Форма создания задания открыта: {task_form_url}")

        # 1. Название задания — с датой и временем, чтобы каждое было уникальным
        task_name = f"Автообзвон {datetime.now().strftime('%d %m %Y %H %M')}"
        print(f"📝 Задаю название: {task_name}")
        # Сохраняем название сразу — пригодится при поиске строки в списке заданий
        try:
            with open(LAST_TASK_NAME_FILE, "w", encoding="utf-8") as f:
                f.write(task_name)
        except Exception:
            pass
        try:
            name_field = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((
                    By.XPATH, "//label[contains(text(),'Название')]/following::input[1]"
                ))
            )
        except Exception:
            # Запасной вариант: первое текстовое поле на форме
            name_field = driver.find_element(By.CSS_SELECTOR, "input[type='text']")
        name_field.clear()
        name_field.send_keys(task_name)
        time.sleep(0.3)

        # 2. Выбираем "Синтез речи" для сообщения
        print("🔊 Выбираю «Синтез речи»...")
        synth_tab = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((
                By.XPATH, "//*[self::button or self::div or self::span][contains(text(),'Синтез речи')]"
            ))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", synth_tab)
        time.sleep(0.3)
        driver.execute_script("arguments[0].click();", synth_tab)
        time.sleep(1)

        # Проверяем что вкладка реально переключилась (должна появиться кнопка "Создать сообщение")
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//button[contains(., 'Создать сообщение')]"))
            )
            print("   ✅ Вкладка «Синтез речи» активна")
        except Exception:
            print("   ⚠️  Похоже вкладка не переключилась, пробую кликнуть ещё раз...")
            driver.execute_script("arguments[0].click();", synth_tab)
            time.sleep(1)

        # 3. Нажимаем "Создать сообщение"
        create_msg_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Создать сообщение')]"))
        )
        try:
            create_msg_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", create_msg_btn)
        time.sleep(1)

        # 4. Вводим текст сообщения в открывшемся модальном окне
        print("✍️  Ввожу текст сообщения...")
        text_area = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Введите текст')]/following::textarea[1]"))
        )
        text_area.clear()
        text_area.send_keys(MTS_MESSAGE_TEXT)
        time.sleep(0.5)

        # 5. Нажимаем "Синтезировать речь"
        synth_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Синтезировать речь')]"))
        )
        try:
            synth_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", synth_btn)
        print("⏳ Жду синтеза речи (до 20 сек)...")
        time.sleep(5)

        # 6. Нажимаем "Применить"
        apply_btn = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Применить')]"))
        )
        try:
            apply_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", apply_btn)
        print("   ✅ Сообщение применено")
        time.sleep(1)

        # 7. Включаем "Собирать ответы"
        print("🔘 Включаю сбор ответов...")
        try:
            collect_toggle_label = driver.find_element(
                By.XPATH, "//*[contains(text(),'Собирать ответы')]/following::label[contains(@class,'switch')][1]"
            )
            try:
                collect_toggle_label.click()
            except Exception:
                driver.execute_script("arguments[0].click();", collect_toggle_label)
        except Exception:
            print("   ⚠️  Не нашёл переключатель «Собирать ответы» — проверь вручную.")
        time.sleep(1)

        # 8. Выбираем опрос "Итог" в выпадающем списке настройки сбора ответов
        print(f"📋 Выбираю опрос «{MTS_SURVEY_NAME}»...")
        survey_selected = False

        # Стратегия 1: найти поле ввода именно в пределах блока "Настройка сбора ответов"
        # (ищем ближайший общий контейнер, а не просто "первый input после текста")
        try:
            container = driver.find_element(
                By.XPATH, "//*[contains(text(),'Настройка сбора ответов')]/ancestor::*[self::div][1]/.."
            )
            dropdown_input = container.find_element(By.TAG_NAME, "input")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown_input)
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", dropdown_input)
            time.sleep(0.7)

            survey_option = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, f"//*[normalize-space(text())='{MTS_SURVEY_NAME}']"))
            )
            driver.execute_script("arguments[0].click();", survey_option)
            survey_selected = True
        except Exception:
            pass

        # Стратегия 2: запасной вариант — просто кликнуть на видимый текст "Итог" на странице,
        # если он уже отображается (например выпадающий список открылся сам или это единственная опция)
        if not survey_selected:
            try:
                survey_option = driver.find_element(
                    By.XPATH, f"//*[normalize-space(text())='{MTS_SURVEY_NAME}']"
                )
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", survey_option)
                driver.execute_script("arguments[0].click();", survey_option)
                survey_selected = True
            except Exception:
                pass

        if survey_selected:
            print(f"   ✅ Опрос «{MTS_SURVEY_NAME}» выбран")
        else:
            print(f"   ⚠️  Не удалось выбрать опрос автоматически — выбери «{MTS_SURVEY_NAME}» вручную.")
            input("   -> Нажми Enter после выбора опроса вручную: ")
        time.sleep(1)

        # 9. Загружаем файл номеров
        print(f"📤 Загружаю файл номеров: {mts_xlsx_abspath}")
        file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file'].file-input")
        file_input.send_keys(mts_xlsx_abspath)

        # 10. Подтверждаем модальное окно "Добавление номеров".
        # ВАЖНО: на странице ДВЕ кнопки с текстом "Загрузить список" —
        # одна обычная (открывает системный диалог выбора файла), и одна
        # внутри модального окна (подтверждает уже выбранный файл). Ищем
        # именно вторую — она находится рядом с текстом "будет добавлен".
        print("⏳ Жду модальное окно подтверждения номеров (опрашиваю каждые 0.5 сек, до 30 сек)...")
        confirmed = False
        deadline = time.time() + 30

        while time.time() < deadline and not confirmed:
            try:
                # Сначала находим сам текст модального окна — он появляется ТОЛЬКО
                # когда окно подтверждения открыто
                modal_marker = driver.find_element(By.XPATH, "//*[contains(text(),'будет добавлен')]")
                if modal_marker.is_displayed():
                    print(f"   Модальное окно найдено: «{modal_marker.text}»")

                    # Кнопка подтверждения находится в том же диалоговом контейнере,
                    # что и этот текст — ищем её через общего предка
                    confirm_btn = driver.find_element(
                        By.XPATH,
                        "//*[contains(text(),'будет добавлен')]/ancestor::*[self::div][3]//button[contains(., 'Загрузить список')]"
                    )

                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", confirm_btn)
                    time.sleep(0.3)
                    driver.execute_script("arguments[0].click();", confirm_btn)
                    time.sleep(1)

                    # Проверяем что модальное окно реально закрылось (текст исчез)
                    try:
                        still_there = driver.find_element(By.XPATH, "//*[contains(text(),'будет добавлен')]")
                        if still_there.is_displayed():
                            continue  # окно всё ещё открыто, клик не сработал — пробуем снова
                    except Exception:
                        pass  # текст пропал — окно закрылось, успех

                    confirmed = True
                    print("   ✅ Подтвердил загрузку списка номеров")
            except Exception:
                pass

            if not confirmed:
                time.sleep(0.5)

        if not confirmed:
            print("   ⚠️  Модальное окно подтверждения не появилось или не найдено.")
            print("   ⚠️  ОБЯЗАТЕЛЬНО посмотри на экран браузера прямо сейчас:")
            print("   ⚠️  если видишь окно 'Добавление номеров' — нажми в нём")
            print("   ⚠️  кнопку 'Загрузить список' САМ, и только потом жми Enter здесь.")
            input("   -> Нажми Enter здесь ПОСЛЕ клика в браузере (или если окна не было): ")

        # 10.5 Настраиваем расписание: Пн-Пт, время 07:00 - 21:00
        print("📅 Настраиваю расписание (Пн-Пт, 07:00-21:00)...")
        try:
            schedule_link = driver.find_element(By.XPATH, "//*[contains(text(),'Настроить расписание')]")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", schedule_link)
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", schedule_link)
            time.sleep(1)

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Настройка расписания автоинформирования')]"))
            )

            # Находим модальное окно расписания целиком, чтобы искать поля
            # именно внутри него (а не по всей странице)
            modal = driver.find_element(
                By.XPATH, "//*[contains(text(),'Настройка расписания автоинформирования')]/ancestor::*[self::div][3]"
            )

            # Все текстовые поля времени внутри модального окна идут по порядку:
            # [Пн-с, Пн-по, Вт-с, Вт-по, Ср-с, Ср-по, Чт-с, Чт-по, Пт-с, Пт-по, ...]
            all_time_fields = modal.find_elements(By.CSS_SELECTOR, "input[type='text'], input:not([type])")
            print(f"   Найдено полей времени в окне: {len(all_time_fields)}")

            days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
            for i, day in enumerate(days):
                start_idx = i * 2
                if start_idx + 1 >= len(all_time_fields):
                    print(f"   ⚠️  Не хватает полей для «{day}»")
                    continue

                field_from = all_time_fields[start_idx]
                field_to = all_time_fields[start_idx + 1]

                for field, value in [(field_from, "07:00"), (field_to, "21:00")]:
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", field)
                    driver.execute_script("arguments[0].click();", field)
                    field.send_keys(Keys.CONTROL, "a")
                    field.send_keys(value)
                    field.send_keys(Keys.TAB)
                    time.sleep(0.2)
                print(f"   {day}: 07:00-21:00")

            # ВАЖНО: ищем кнопку "Сохранить" ТОЛЬКО внутри модального окна расписания
            # (переменная modal, найденная чуть выше), а не по всей странице. Раньше
            # поиск шёл по всему документу и брал "последнюю видимую" кнопку с текстом
            # "Сохранить" — но is_displayed() не учитывает перекрытие оверлеем, поэтому
            # под модалкой основная кнопка "Сохранить задание" тоже считалась "видимой",
            # и иногда клик улетал по НЕЙ — это сохраняло и закрывало всё задание сразу,
            # минуя настройку дозвона и включение информирования.
            schedule_save_btn = WebDriverWait(modal, 10).until(
                EC.presence_of_element_located((By.XPATH, ".//button[contains(., 'Сохранить')]"))
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", schedule_save_btn)
            time.sleep(0.2)
            driver.execute_script("arguments[0].click();", schedule_save_btn)
            print("   ✅ Расписание настроено: Пн-Пт, 07:00-21:00")
            time.sleep(1)

            # Подстраховка: проверяем, что модальное окно расписания реально закрылось,
            # а не то, что страница перешла на сохранение всего задания (URL не должен
            # поменяться, и заголовок модалки должен исчезнуть)
            try:
                still_open = driver.find_elements(
                    By.XPATH, "//*[contains(text(),'Настройка расписания автоинформирования')]"
                )
                if any(el.is_displayed() for el in still_open):
                    print("   ⚠️  Модальное окно расписания всё ещё открыто — проверь вручную.")
            except Exception:
                pass

            # КРИТИЧЕСКАЯ ПРОВЕРКА: если после клика URL уехал от формы создания
            # задания — значит клик по ошибке попал по кнопке сохранения ВСЕГО
            # задания (а не модалки расписания), и задание уже сохранено/закрыто
            # без настройки дозвона и включения информирования. Продолжать
            # дальнейшие шаги на этой странице бессмысленно — лучше остановиться
            # сразу с понятной ошибкой, чем тихо наделать что-то ещё не туда.
            if driver.current_url != task_form_url:
                print("\n" + "🛑" * 20)
                print("  ОШИБКА: после сохранения расписания страница изменилась")
                print(f"  (было: {task_form_url}")
                print(f"   стало: {driver.current_url})")
                print("  Похоже, что вместо кнопки модалки расписания нажалась")
                print("  кнопка сохранения ВСЕГО задания, и оно уже сохранено")
                print("  БЕЗ настройки дозвона и БЕЗ включённого информирования!")
                print("  Зайди в МТС, найди это задание и донастрой его вручную:")
                print("   - Максимум попыток / перезвон через")
                print("   - Включить информирование")
                print("🛑" * 20)
                return
        except Exception as e:
            print(f"   ⚠️  Не удалось настроить расписание автоматически ({e}).")
            print("   ⚠️  Настрой вручную: Пн-Пт, время 07:00-21:00, затем нажми Enter.")
            input("   -> Нажми Enter после настройки расписания вручную: ")

        # 10.6 Настройка дозвона: автоматизация этого конкретного виджета
        # оказалась ненадёжной (плавающий React-компонент) — настраиваем вручную.
        print("\n" + "📞" * 20)
        print("  ОСТАНОВКА: настрой параметры дозвона ВРУЧНУЮ на экране браузера:")
        print("  - Максимум попыток: выбери 3")
        print("  - Перезванивать через: впиши 5")
        print("📞" * 20)
        input("  -> Когда настроишь — нажми Enter ЗДЕСЬ (не раньше!): ")
        print("   Продолжаю...")
        time.sleep(0.5)

        # 11. "Включить информирование" — после нескольких неудачных попыток
        # надёжно автоматизировать этот виджет (разметка/поведение не дались
        # под автоклик), оставляем настройку вручную, как и параметры дозвона.
        print("\n" + "🔘" * 20)
        print("  ОСТАНОВКА: включи тумблер «Включить информирование» ВРУЧНУЮ")
        print("  на экране браузера (должен стать синим/зелёным).")
        print("🔘" * 20)
        input("  -> Когда включишь — нажми Enter ЗДЕСЬ: ")
        print("   Продолжаю...")
        time.sleep(0.5)

        # Финальная пауза для визуальной проверки перед сохранением
        print("\n" + "=" * 55)
        print("  ПРОВЕРЬ НА ЭКРАНЕ БРАУЗЕРА:")
        print("  - правильный текст сообщения?")
        print("  - правильное количество номеров?")
        print("  - расписание: Пн-Пт, 07:00-21:00?")
        print("  - максимум попыток: 3, перезвон через: 5 минут?")
        print("  - включён ли тумблер «Включить информирование»?")
        print("=" * 55)
        print("  Если всё верно — нажми Enter, чтобы сохранить и запустить обзвон.")
        print("  Если что-то не так — закрой окно браузера вместо Enter.")
        input("  -> Нажми Enter: ")

        # 12. Сохраняем
        print("💾 Сохраняю задание (это запускает обзвон)...")
        save_btn = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//button[contains(., 'Сохранить')]"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", save_btn)
        time.sleep(0.5)
        try:
            save_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", save_btn)
        time.sleep(3)

        print(f"\n✅ Обзвон запущен!")
        print(f"   Ссылка на задание: {driver.current_url}")
        print(f"\n   Сохрани эту ссылку — она понадобится команде fetch_report.")
        print(f"   Обычно обзвон полностью завершается (с учётом")
        print(f"   перезвонов) примерно за 15 минут.")

        # Сохраняем ссылку в файл, чтобы START.bat мог подхватить её
        # автоматически в пункте «Полный цикл» — без ручного копирования.
        try:
            with open(LAST_TASK_URL_FILE, "w", encoding="utf-8") as f:
                f.write(driver.current_url.strip())
        except Exception as e:
            print(f"   ⚠️  Не удалось сохранить ссылку в файл {LAST_TASK_URL_FILE}: {e}")

        # Возвращаем url для использования в полном цикле
        return driver.current_url

    finally:
        if own_driver:
            driver.quit()
            print("Браузер закрыт.")



def cmd_fetch_report(task_url: str, external_driver=None):
    """
    Команда: открыть указанное задание в МТС, скачать отчёт по нему,
    и сразу запустить объединение с заказами CRM (merge).
    """
    own_driver = external_driver is None
    driver = external_driver if external_driver else create_driver()
    try:
        print(f"🌐 Открываю задание: {task_url}")
        driver.get(task_url)

        time.sleep(2)
        logged_in_automatically = try_mts_login(driver)

        print()
        print("=" * 55)
        if logged_in_automatically:
            print("  Вход выполнен автоматически.")
        else:
            print("  Если МТС попросит войти — авторизуйся в открывшемся окне.")
        print("  Дождись пока полностью загрузится страница задания.")
        print("  Нажми Enter здесь, когда увидишь форму задания.")
        print("=" * 55)
        input("  -> Нажми Enter: ")
        print()

        # Кнопка скачивания отчёта обычно находится в списке заданий, а не в самой
        # форме задания — переходим в список и ищем строку с этим заданием.
        print("📥 Перехожу в список заданий для скачивания отчёта...")
        driver.get(MTS_TASKS_LIST_URL)

        # Явно ждём пока React отрисует строки таблицы заданий (не просто загрузку HTML)
        print("   Жду загрузки таблицы заданий МТС...")
        try:
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//tr[td]"))
            )
            time.sleep(1)  # небольшая пауза после появления строк — React может ещё анимировать
        except Exception:
            time.sleep(4)  # запасной вариант — просто подождать

        # Засекаем список файлов в Downloads ДО скачивания, чтобы потом найти новый файл
        before_files = set(os.listdir(DOWNLOADS_DIR)) if os.path.isdir(DOWNLOADS_DIR) else set()

        print("🔍 Ищу кнопку скачивания отчёта последнего задания...")
        download_clicked = False

        # Определяем по какому признаку искать строку нашего задания в таблице.
        # Приоритет: 1) название задания (самое надёжное — уникальное и текстовое),
        # 2) uuid из URL (если есть ссылка с uuid), 3) первая строка (запасной вариант).
        task_uuid = task_url.rstrip("/").split("/")[-1] if "/tasks/" in task_url else None

        # Читаем сохранённое название задания
        task_name = None
        try:
            if os.path.exists(LAST_TASK_NAME_FILE):
                with open(LAST_TASK_NAME_FILE, encoding="utf-8") as f:
                    task_name = f.read().strip()
        except Exception:
            pass

        # Строим список приоритетных row_prefix для поиска
        row_prefixes = []
        if task_name:
            row_prefixes.append(f"//tr[contains(.,'{task_name}')]")
        if task_uuid:
            row_prefixes.append(f"//tr[.//a[contains(@href,'{task_uuid}')]]")
        # Запасной: первая строка таблицы
        row_prefixes.append("(//tbody/tr)[1]")
        row_prefixes.append("(//tr[td])[1]")

        for row_prefix in row_prefixes:
            for suffix in [
                "//*[contains(@class,'download')]",
                "//button[contains(@title,'Скача') or contains(@aria-label,'Скача') or contains(@title,'отчёт')]",
                "//a[contains(@href,'export') or contains(@href,'download') or contains(@href,'report')]",
                "//i[contains(@class,'download') or contains(@class,'export')]/..",
            ]:
                try:
                    btn = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, row_prefix + suffix))
                    )
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                    time.sleep(0.3)
                    btn.click()
                    download_clicked = True
                    print("   ✅ Кнопка скачивания найдена и нажата")
                    break
                except Exception:
                    continue
            if download_clicked:
                break

        if not download_clicked:
            print()
            print("   ⚠️  Не нашёл кнопку скачивания автоматически.")
            print("   Сделай это вручную:")
            print("   1. Посмотри на открытый браузер — там список заданий МТС")
            print("   2. Нажми кнопку скачивания отчёта (иконка ↓) у нужного задания")
            print("   3. В появившемся окне выбери CSV и нажми «Скачать»")
            print("   4. Дождись скачивания и вернись сюда")
            input("   -> Нажми Enter после скачивания: ")

        # После клика по ↓ МТС показывает модальное окно выбора формата
        # (Excel / CSV) с кнопкой «Скачать» — всё на той же странице, без новой вкладки.
        # Выбираем CSV и жмём «Скачать».
        print("   Жду модальное окно выбора формата...")
        try:
            # Ждём появления кнопок выбора формата
            csv_btn = WebDriverWait(driver, 8).until(
                EC.element_to_be_clickable((By.XPATH,
                    "//button[normalize-space(.)='CSV'] | //label[normalize-space(.)='CSV']"
                ))
            )
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", csv_btn)
            time.sleep(0.2)
            csv_btn.click()
            print("   ✅ Выбран формат CSV")
            time.sleep(0.5)

            confirm_btn = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH,
                    "//button[contains(normalize-space(.),'Скачать')]"
                ))
            )
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", confirm_btn)
            time.sleep(0.2)
            confirm_btn.click()
            print("   ✅ Нажата кнопка «Скачать»")
        except Exception as e:
            print(f"   ⚠️  Модальное окно не появилось или формат уже выбран: {e}")

        # Ждём появления нового файла в Downloads
        # Важно: Chrome сначала создаёт файл с расширением .tmp (.crdownload),
        # и только после завершения переименовывает в настоящий файл.
        # Ждём пока появится готовый файл (не .tmp и не .crdownload).
        print("⏳ Жду скачивания файла...")
        new_file = None
        for _ in range(60):
            time.sleep(1)
            if os.path.isdir(DOWNLOADS_DIR):
                after_files = set(os.listdir(DOWNLOADS_DIR))
                diff = after_files - before_files
                # Игнорируем временные файлы Chrome
                diff = {f for f in diff if not f.endswith(".crdownload") and not f.endswith(".tmp")}
                if diff:
                    new_file = sorted(diff, key=lambda f: os.path.getmtime(
                        os.path.join(DOWNLOADS_DIR, f)))[-1]
                    break

        if not new_file:
            print("❌ Не удалось автоматически найти скачанный файл.")
            print(f"   Проверь папку {DOWNLOADS_DIR} вручную и запусти merge с этим файлом:")
            print(f"   python scrape_crm.py merge ИМЯ_ФАЙЛА")
            return

        report_path = os.path.join(DOWNLOADS_DIR, new_file)
        print(f"✅ Отчёт скачан: {report_path}")

        # Копируем отчёт в рабочую папку скрипта для удобства
        local_copy = os.path.basename(report_path)
        try:
            with open(report_path, "rb") as src, open(local_copy, "wb") as dst:
                dst.write(src.read())
            print(f"   Копия сохранена рядом со скриптом: {local_copy}")
            # Удаляем оригинал из Downloads — он уже скопирован в C:\obzvon
            try:
                os.remove(report_path)
            except Exception:
                pass
            report_path = local_copy
        except Exception:
            pass

    finally:
        if own_driver:
            driver.quit()
            print("Браузер закрыт.")

    # Сразу запускаем объединение с заказами CRM
    print("\n🔗 Запускаю объединение с заказами CRM...")
    cmd_merge(report_path)



def cmd_full_cycle():
    """
    Полный цикл шагов 2+3 в одном браузере — создание задания МТС
    и скачивание отчёта без повторной авторизации.
    """
    driver = create_driver()
    try:
        task_url = cmd_call_new(external_driver=driver)
        if not task_url:
            print("❌ Не удалось получить ссылку на задание, скачивание пропущено.")
            return
        cmd_fetch_report(task_url, external_driver=driver)
    finally:
        driver.quit()
        print("Браузер закрыт.")

# ============================================================
#  ТОЧКА ВХОДА
# ============================================================

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    command = sys.argv[1]

    if command == "crm":
        cmd_crm()
    elif command == "call":
        if len(sys.argv) > 2 and sys.argv[2] == "--copy":
            resume_url = sys.argv[3] if len(sys.argv) > 3 else None
            cmd_call_copy(resume_url)
        else:
            cmd_call_new()
    elif command == "full_cycle":
        cmd_full_cycle()
    elif command == "fetch_report":
        if len(sys.argv) < 3:
            print("Укажи ссылку на задание МТС:")
            print("  python scrape_crm.py fetch_report https://.../autocaller/tasks/...")
            sys.exit(1)
        cmd_fetch_report(sys.argv[2])
    elif command == "merge":
        if len(sys.argv) < 3:
            print("Укажи путь к файлу отчёта МТС:")
            print("  python scrape_crm.py merge ОТЧЕТ.csv")
            sys.exit(1)
        cmd_merge(sys.argv[2])
    else:
        print(f"Неизвестная команда: {command}")
        print(__doc__)
        sys.exit(1)


if __name__ == "__main__":
    main()
